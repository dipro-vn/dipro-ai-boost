#!/usr/bin/env python3
"""
Backlog Statistics Report Generator
Tự động hóa báo cáo Backlog và xuất Excel.

Usage:
    python backlog_report.py
    python backlog_report.py --from-date 2025-01-01 --to-date 2025-01-31
    python backlog_report.py --from-date 2025-01-01 --to-date 2025-01-31 --output my_report.xlsx
    python backlog_report.py --from-date 2026-03-01 --to-date 2026-04-30 --plan-new-month
"""

import os
import sys
import json
import logging
import argparse
from datetime import datetime, timedelta
from typing import Dict, List, Optional

import requests
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURATION
# ============================================================

# Member role mapping (data/member.json)
_MEMBER_JSON_PATH = os.path.join(os.path.dirname(__file__), "data", "member.json")
try:
    with open(_MEMBER_JSON_PATH, encoding="utf-8") as _f:
        MEMBER_ROLES: Dict[str, str] = json.load(_f)
except FileNotFoundError:
    MEMBER_ROLES = {}

WORKING_HOURS_PER_DAY = 7
MAX_ISSUES_PER_REQUEST = 100
AVAILABILITY_DAYS_AHEAD = 14

# Local config (local.json — không commit, chứa các key riêng của môi trường)
_LOCAL_JSON_PATH = os.path.join(os.path.dirname(__file__), "local.json")
try:
    with open(_LOCAL_JSON_PATH, encoding="utf-8") as _lf:
        _LOCAL_CFG: Dict = json.load(_lf)
except FileNotFoundError:
    _LOCAL_CFG = {}


def _require(key: str):
    val = _LOCAL_CFG.get(key) or os.environ.get(key.upper())
    if not val:
        sys.stderr.write(
            f"❌ Thiếu '{key}' trong local.json (hoặc env {key.upper()}).\n"
            f"   Xem local.json.example để biết cấu trúc.\n"
        )
        sys.exit(1)
    return val


BACKLOG_BASE_URL: str = _require("backlog_base_url").rstrip("/")
API_KEY:          str = _require("api_key")
PROJECT_ID:       int = int(_require("project_id"))

# Base URL cho web link (bỏ '/api/v2' → chỉ giữ host)
BACKLOG_WEB_URL:  str = BACKLOG_BASE_URL.rsplit("/api/", 1)[0]

# Team groups dùng cho plan resource
_PLAN_TEAM_GROUPS: Dict[str, List[str]] = {
    "BE":       ["BE"],
    "FE":       ["FE"],
    "Mobile":   ["Mobile"],
    "QC":       ["QC"],
    "Designer": ["Designer"],
    "INFRA":    ["INFRA"],
}
_PLAN_ROLE_TO_TEAM: Dict[str, str] = {
    role: team
    for team, roles in _PLAN_TEAM_GROUPS.items()
    for role in roles
}

# Statuses được coi là "hoàn thành" (không tính là overdue)
DONE_STATUSES = {"Resolved", "Closed", "Done", "完了", "解決済み", "処理済み"}

# Members bị loại khỏi toàn bộ thống kê — đọc từ data/excludeds.txt (mỗi dòng 1 tên)
_EXCLUDEDS_PATH = os.path.join(os.path.dirname(__file__), "data", "excludeds.txt")
try:
    with open(_EXCLUDEDS_PATH, encoding="utf-8") as _ef:
        EXCLUDED_MEMBERS: set = {line.strip() for line in _ef if line.strip()}
except FileNotFoundError:
    EXCLUDED_MEMBERS = set()

# Kế hoạch nhân sự từ file (dùng cho --plan-new-month)
_PLAN_RESOURCE_JSON_PATH = os.path.join(os.path.dirname(__file__), "data", "plan_resource.json")
try:
    with open(_PLAN_RESOURCE_JSON_PATH, encoding="utf-8") as _pf:
        PLAN_RESOURCE: List[Dict] = json.load(_pf)
except FileNotFoundError:
    PLAN_RESOURCE = []

# Member allocation mapping: Backlog User ID (int) → ratio
# Đọc từ plan_resource.json, field "backlog_id" + "plan"
# 1.0 = 100% = 40h/tuần, 0.5 = 50% = 20h/tuần
# ID lấy từ cột assigneeId trong Raw_Data hoặc GET /api/v2/users
MEMBER_ALLOCATION_BY_ID: Dict[int, float] = {
    int(e["backlog_id"]): float(e.get("plan", 1.0))
    for e in PLAN_RESOURCE
    if e.get("backlog_id")
}

# Fallback: khi backlog_id chưa fill (=0), match bằng tên account/full_name
MEMBER_ALLOCATION_BY_NAME: Dict[str, float] = {
    (e.get("account") or e.get("full_name") or "").strip(): float(e.get("plan", 1.0))
    for e in PLAN_RESOURCE
    if (e.get("account") or e.get("full_name"))
}

_NEED_MD_PATH = os.path.join(os.path.dirname(__file__), "data", "need.md")
try:
    with open(_NEED_MD_PATH, encoding="utf-8") as _nf:
        NEED_EXTRA: str = _nf.read().strip()
except FileNotFoundError:
    NEED_EXTRA = ""

# Logging setup
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


# ============================================================
# 1. EXTRACT — Lấy dữ liệu từ Backlog API
# ============================================================

def get_backlog_data(from_date: str, to_date: str) -> pd.DataFrame:
    """
    Trích xuất toàn bộ Issues từ Backlog API với xử lý phân trang.

    Args:
        from_date: Ngày bắt đầu (YYYY-MM-DD)
        to_date  : Ngày kết thúc (YYYY-MM-DD)

    Returns:
        DataFrame chứa toàn bộ issues đã được parse sạch
    """
    logger.info(f"Fetching Backlog issues với dueDate [{from_date} → {to_date}]...")

    all_issues: List[dict] = []
    offset = 0

    while True:
        params = {
            "apiKey": API_KEY,
            "projectId[]": PROJECT_ID,
            "count": MAX_ISSUES_PER_REQUEST,
            "offset": offset,
            "dueDateSince": from_date,
            "dueDateUntil": to_date,
            "order": "dueDate",
        }

        try:
            resp = requests.get(
                f"{BACKLOG_BASE_URL}/issues",
                params=params,
                timeout=30,
            )
            resp.raise_for_status()
            batch: List[dict] = resp.json()

        except requests.exceptions.Timeout:
            logger.error("Backlog API timeout — please retry.")
            raise
        except requests.exceptions.HTTPError as exc:
            logger.error(f"HTTP {exc.response.status_code}: {exc.response.text[:300]}")
            raise
        except requests.exceptions.RequestException as exc:
            logger.error(f"Request failed: {exc}")
            raise

        if not batch:
            break

        all_issues.extend(batch)
        logger.info(f"  → {len(all_issues)} issues fetched (offset={offset})")

        if len(batch) < MAX_ISSUES_PER_REQUEST:
            break

        offset += MAX_ISSUES_PER_REQUEST

    logger.info(f"Total issues fetched (before filter): {len(all_issues)}")

    if not all_issues:
        logger.warning("No issues found for the given date range.")
        return pd.DataFrame()

    df = _parse_issues(all_issues)

    # Filter theo dueDate trong khoảng [from_date, to_date] (API đã lọc dueDateSince/Until nhưng
    # trả về UTC nên cần re-check sau khi convert sang giờ VN)
    from_naive = pd.Timestamp(from_date + " 00:00:00")
    to_naive   = pd.Timestamp(to_date   + " 23:59:59")
    df = df[(df["dueDate"] >= from_naive) & (df["dueDate"] <= to_naive)].reset_index(drop=True)
    logger.info(f"Issues after dueDate filter [{from_date} → {to_date}]: {len(df)}")

    # Loại bỏ các member trong danh sách EXCLUDED_MEMBERS
    if EXCLUDED_MEMBERS and "assignee" in df.columns:
        before = len(df)
        df = df[~df["assignee"].isin(EXCLUDED_MEMBERS)].reset_index(drop=True)
        removed = before - len(df)
        if removed:
            logger.info(f"Excluded {removed} issues from members: {EXCLUDED_MEMBERS}")

    return df


def _fetch_backlog_users_by_email() -> Dict[str, Dict]:
    """Fetch danh sách user Backlog, trả về mapping email(lower) → user dict.
    Trả về {} nếu API lỗi (không raise)."""
    try:
        resp = requests.get(
            f"{BACKLOG_BASE_URL}/users",
            params={"apiKey": API_KEY},
            timeout=15,
        )
        resp.raise_for_status()
        users = resp.json()
    except requests.exceptions.RequestException as exc:
        logger.warning(f"  [plan_resource] không lấy được danh sách user Backlog: {exc}")
        return {}
    return {
        (u.get("mailAddress") or "").strip().lower(): u
        for u in users
        if u.get("mailAddress")
    }


def _autofill_plan_resource_ids(df: pd.DataFrame) -> None:
    """
    Điền backlog_id còn trống (0) trong data/plan_resource.json.
    Ưu tiên match theo email (nếu có), fallback match theo tên assignee.
    Cập nhật MEMBER_ALLOCATION_BY_ID in-place.
    """
    if not PLAN_RESOURCE:
        return

    # Build mapping tên → id từ dữ liệu Backlog (fallback khi thiếu email)
    name_to_id: Dict[str, int] = {}
    if not df.empty:
        for _, row in df[df["assigneeId"].notna()][["assignee", "assigneeId"]].drop_duplicates().iterrows():
            name_to_id[str(row["assignee"]).strip()] = int(row["assigneeId"])

    # Chỉ fetch user API 1 lần nếu có entry cần match theo email
    email_to_user: Optional[Dict[str, Dict]] = None
    def _email_map() -> Dict[str, Dict]:
        nonlocal email_to_user
        if email_to_user is None:
            email_to_user = _fetch_backlog_users_by_email()
        return email_to_user

    changed = False
    for entry in PLAN_RESOURCE:
        if entry.get("backlog_id"):
            continue

        email = (entry.get("email") or "").strip().lower()
        if email:
            user = _email_map().get(email)
            if user:
                uid = int(user["id"])
                entry["backlog_id"] = uid
                if not entry.get("account"):
                    entry["account"] = user.get("userId", "")
                if not entry.get("full_name"):
                    entry["full_name"] = user.get("name", "")
                MEMBER_ALLOCATION_BY_ID[uid] = float(entry.get("plan", 1.0))
                changed = True
                logger.info(f"  [plan_resource] auto-fill via email: {email} → backlog_id={uid}")
                continue

        name = (entry.get("account") or entry.get("full_name") or "").strip()
        uid = name_to_id.get(name)
        if uid:
            entry["backlog_id"] = uid
            MEMBER_ALLOCATION_BY_ID[uid] = float(entry.get("plan", 1.0))
            changed = True
            logger.info(f"  [plan_resource] auto-fill via name: {name} → backlog_id={uid}")

    if changed:
        try:
            tmp = _PLAN_RESOURCE_JSON_PATH + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(PLAN_RESOURCE, f, ensure_ascii=False, indent=2)
            os.replace(tmp, _PLAN_RESOURCE_JSON_PATH)
        except OSError as exc:
            logger.warning(f"Không ghi được plan_resource.json: {exc}")


def get_project_key(default: str = "report") -> str:
    """Lấy projectKey (ví dụ: ESKITCHEN) từ Backlog để đặt tên file output.
    Trả về `default.lower()` nếu API lỗi hoặc network chưa sẵn sàng."""
    try:
        resp = requests.get(
            f"{BACKLOG_BASE_URL}/projects/{PROJECT_ID}",
            params={"apiKey": API_KEY},
            timeout=15,
        )
        resp.raise_for_status()
        key = (resp.json() or {}).get("projectKey") or default
        return key.lower()
    except requests.exceptions.RequestException as exc:
        logger.warning(f"Không lấy được projectKey ({exc}); dùng '{default}'")
        return default.lower()


def get_full_month_data(from_date: str) -> pd.DataFrame:
    """Fetch tất cả issues có dueDate trong tháng của from_date (dùng cho Monthly_Stats)."""
    month_start = pd.Timestamp(from_date).to_period("M").to_timestamp()
    month_end   = month_start + pd.offsets.MonthEnd(0)
    from_str    = month_start.strftime("%Y-%m-%d")
    to_str      = month_end.strftime("%Y-%m-%d")

    logger.info(f"Fetching full month [{from_str} → {to_str}] for Monthly_Stats...")
    all_issues: List[dict] = []
    offset = 0
    while True:
        params = {
            "apiKey"       : API_KEY,
            "projectId[]"  : PROJECT_ID,
            "count"        : MAX_ISSUES_PER_REQUEST,
            "offset"       : offset,
            "dueDateSince" : from_str,
            "dueDateUntil" : to_str,
        }
        try:
            resp = requests.get(f"{BACKLOG_BASE_URL}/issues", params=params, timeout=30)
            resp.raise_for_status()
            batch: List[dict] = resp.json()
        except requests.exceptions.RequestException as exc:
            logger.warning(f"Monthly fetch failed: {exc}")
            return pd.DataFrame()
        if not batch:
            break
        all_issues.extend(batch)
        if len(batch) < MAX_ISSUES_PER_REQUEST:
            break
        offset += MAX_ISSUES_PER_REQUEST

    logger.info(f"  → {len(all_issues)} issues in full month")
    if not all_issues:
        return pd.DataFrame()

    df = _parse_issues(all_issues)
    from_naive = pd.Timestamp(from_str)
    to_naive   = pd.Timestamp(to_str + " 23:59:59")
    df = df[(df["dueDate"] >= from_naive) & (df["dueDate"] <= to_naive)].reset_index(drop=True)
    if EXCLUDED_MEMBERS and "assignee" in df.columns:
        df = df[~df["assignee"].isin(EXCLUDED_MEMBERS)].reset_index(drop=True)
    return df


def _parse_issues(raw: List[dict]) -> pd.DataFrame:
    """Chuyển đổi raw JSON từ Backlog API thành DataFrame sạch."""
    records = []

    for issue in raw:
        assignee   = issue.get("assignee") or {}
        status     = issue.get("status")   or {}
        issue_type = issue.get("issueType") or {}
        priority   = issue.get("priority") or {}
        categories = issue.get("category")  or []
        milestones = issue.get("milestone") or []

        records.append({
            "id"            : issue.get("id"),
            "issueKey"      : issue.get("issueKey", ""),
            "summary"       : issue.get("summary", ""),
            "status"        : status.get("name", "Unknown"),
            "priority"      : priority.get("name", ""),
            "issueType"     : issue_type.get("name", ""),
            "assignee"      : assignee.get("name") or "Unassigned",
            "assigneeId"    : assignee.get("id"),
            "category"      : ", ".join(c.get("name", "") for c in categories) or "None",
            "milestone"     : ", ".join(m.get("name", "") for m in milestones) or "None",
            "estimatedHours": issue.get("estimatedHours"),
            "actualHours"   : issue.get("actualHours"),
            "startDate"     : issue.get("startDate"),
            "dueDate"       : issue.get("dueDate"),
            "created"       : issue.get("created"),
            "updated"       : issue.get("updated"),
            "description"   : (issue.get("description") or "")[:500],
            "backlogUrl"    : f"{BACKLOG_WEB_URL}/view/{issue.get('issueKey', '')}",
        })

    df = pd.DataFrame(records)

    # Parse dates — Backlog trả về UTC, convert sang giờ Việt Nam (UTC+7)
    for col in ("startDate", "dueDate", "created", "updated"):
        df[col] = (
            pd.to_datetime(df[col], errors="coerce", utc=True)
            .dt.tz_convert("Asia/Ho_Chi_Minh")
            .dt.tz_localize(None)
        )

    # Normalize giờ
    df["estimatedHours"] = pd.to_numeric(df["estimatedHours"], errors="coerce").fillna(0)
    df["actualHours"]    = pd.to_numeric(df["actualHours"],    errors="coerce").fillna(0)

    today = pd.Timestamp.now().normalize()

    # Cột tham chiếu ngày: ưu tiên dueDate, fallback updated
    df["refDate"] = df["dueDate"].fillna(df["updated"])

    iso = df["refDate"].dt.isocalendar()
    df["isoWeek"] = iso.week.astype("Int64")
    df["isoYear"] = iso.year.astype("Int64")
    df["yearWeek"] = (
        df["isoYear"].astype(str) + "-W" + df["isoWeek"].astype(str).str.zfill(2)
    )
    df["month"] = df["refDate"].dt.to_period("M").astype(str)

    # Overdue: có dueDate đã qua, chưa hoàn thành
    df["isOverdue"] = (
        df["dueDate"].notna()
        & (df["dueDate"] < today)
        & (~df["status"].isin(DONE_STATUSES))
    )

    # ── Flow Metric: Cycle Time (ngày) ──────────────────────
    # Tính từ startDate (hoặc created nếu không có) đến updated
    start_ref = df["startDate"].fillna(df["created"])
    df["cycleTime"] = (df["updated"] - start_ref).dt.days.clip(lower=0)

    # ── Quality Metric: Estimation Variance ─────────────────
    # > 0: làm lâu hơn dự kiến | < 0: làm nhanh hơn
    df["estimationVariance"] = (df["actualHours"] - df["estimatedHours"]).round(1)

    # ── Type / Priority flags (dùng cho weekly grouping) ────
    bug_pattern = r"Bug|Sửa lỗi|バグ|不具合"
    df["isBug"] = df["issueType"].str.contains(bug_pattern, case=False, na=False)

    blocked_pattern = r"Pending|Wait|Blocked|ブロック|保留|待機"
    df["isBlocked"] = df["status"].str.contains(blocked_pattern, case=False, na=False)

    # ── Predictive Metric: Risk Score ───────────────────────
    # High priority = +2, Overdue = +3, No estimation = +1
    df["riskScore"] = (
        (df["priority"].str.lower() == "high").astype(int) * 2
        + df["isOverdue"].astype(int) * 3
        + (df["estimatedHours"] == 0).astype(int) * 1
    )

    return df


# ============================================================
# 2. PROCESSING — Thống kê với pandas
# ============================================================

def process_stats(df: pd.DataFrame, from_date: str, to_date: str, month_df: Optional[pd.DataFrame] = None) -> Dict:
    """
    Tính toán thống kê tuần, tháng và lịch rảnh của member.

    Returns:
        dict với keys: weekly, monthly, availability, summary, overdue
    """
    if df.empty:
        logger.warning("DataFrame trống — bỏ qua bước thống kê.")
        return {}

    logger.info("Processing statistics...")

    today = pd.Timestamp.now().normalize()

    def _week_key(dt: pd.Timestamp) -> str:
        iso = dt.isocalendar()
        return f"{iso[0]}-W{str(iso[1]).zfill(2)}"

    prev_week_key = _week_key(today - timedelta(weeks=1))
    curr_week_key = _week_key(today)
    next_week_key = _week_key(today + timedelta(weeks=1))

    logger.info(f"  Weeks → prev={prev_week_key}  curr={curr_week_key}  next={next_week_key}")

    from_ts = pd.Timestamp(from_date)
    to_ts   = pd.Timestamp(to_date)

    dashboard_df    = _compute_dashboard_weekly(df, from_ts, to_ts)
    monthly_df      = _compute_monthly_stats(month_df if month_df is not None else df, from_date)
    availability_df = _compute_member_availability(df, from_ts, to_ts)

    # ── Debug dump: tasks dùng để tính availability ──────────
    _debug_availability(df, today)

    action_req_df     = _get_action_required_tasks(df)

    logger.info(f"  Tasks cần chú ý (Action Required): {len(action_req_df)}")
    if not action_req_df.empty:
        for warn_type, grp in action_req_df.groupby("Loại cảnh báo"):
            logger.info(f"    → {warn_type}: {len(grp)} tasks")

    return {
        "dashboard"      : dashboard_df,
        "monthly"        : monthly_df,
        "availability"   : availability_df,
        "action_required": action_req_df,
        "period"         : (from_date, to_date),
    }


def _week_label(week_key: str) -> str:
    """Chuyển 'YYYY-Www' thành 'Tuần mm/dd ~ mm/dd'."""
    year, week = int(week_key[:4]), int(week_key[6:])
    monday = pd.Timestamp.fromisocalendar(year, week, 1)
    sunday = monday + timedelta(days=6)
    return f"Tuần {monday.strftime('%m/%d')} ~ {sunday.strftime('%m/%d')}"


def _alloc_pct(assignee_id, assignee_name: str = "") -> int:
    """Allocation (%) — thử match theo ID trước, fallback theo tên. Mặc định 100%."""
    if pd.notna(assignee_id):
        ratio = MEMBER_ALLOCATION_BY_ID.get(int(assignee_id))
        if ratio is not None:
            return int(round(ratio * 100))
    if assignee_name:
        ratio = MEMBER_ALLOCATION_BY_NAME.get(assignee_name.strip())
        if ratio is not None:
            return int(round(ratio * 100))
    return 100


def _productivity(actual: float, estimate: float) -> Optional[float]:
    """Productivity = Estimate/Actual × 100 (%). >100 tốt (làm nhanh hơn estimate). None nếu chưa log actual."""
    if actual <= 0:
        return None
    return round(estimate / actual * 100, 1)


WEEKLY_CAPACITY_HOURS  = 40   # 100% allocation ≈ 40h/tuần — chuẩn effort load cho Dashboard
EFFORT_LOWER_BAND      = 0.875  # <87.5% (=35h với 100% alloc) → còn dư
EFFORT_UPPER_BAND      = 1.0    # >100% → quá tải


def _effort_label(estimate: float, alloc_pct: int) -> str:
    """So sánh Estimate vs Allocation-hours (100% = 40h/tuần). Trả về 'Còn dư | Đủ | Quá tải'."""
    alloc_hours = alloc_pct / 100 * WEEKLY_CAPACITY_HOURS
    if alloc_hours <= 0:
        return ""
    ratio = estimate / alloc_hours
    if ratio > EFFORT_UPPER_BAND:
        return "Quá tải"
    if ratio >= EFFORT_LOWER_BAND:
        return "Đủ"
    return "Còn dư"


def _compute_dashboard_weekly(
    df: pd.DataFrame, from_ts: pd.Timestamp, to_ts: pd.Timestamp
) -> pd.DataFrame:
    """
    Bảng Dashboard theo tuần — mỗi dòng = 1 member.

    Filter: dueDate ∈ [from_ts, to_ts]. Tổng gộp tất cả issue (task bao gồm bug).

    Columns:
      Member | Role | Allocation (%)
      Total Task | Task Done | Task Remain
      Total Bug  | Bug Done  | Bug Remain
      Total Estimate (h) | Total Actual (h) | Productivity (%) | Effort Load
    """
    if df.empty:
        return pd.DataFrame()

    from_naive = pd.Timestamp(from_ts.date())
    to_naive   = pd.Timestamp(to_ts.date()) + timedelta(days=1) - timedelta(seconds=1)
    sub = df[
        df["dueDate"].notna()
        & (df["dueDate"] >= from_naive) & (df["dueDate"] <= to_naive)
        & (df["assignee"] != "Unassigned")
    ].copy()
    if sub.empty:
        return pd.DataFrame()

    sub["_done"] = sub["status"].isin(DONE_STATUSES).astype(int)
    sub["_is_bug"] = (sub["issueType"].str.strip().str.lower() == "bug").astype(int)
    sub["_bug_done"] = (sub["_is_bug"] & sub["_done"]).astype(int)

    g = (
        sub.groupby("assignee")
        .agg(
            assignee_id     =("assigneeId",     "first"),
            total_task      =("id",             "count"),
            task_done       =("_done",          "sum"),
            total_bug       =("_is_bug",        "sum"),
            bug_done        =("_bug_done",      "sum"),
            total_estimate  =("estimatedHours", "sum"),
            total_actual    =("actualHours",    "sum"),
        )
        .reset_index()
    )

    g["Role"] = g["assignee"].map(lambda a: MEMBER_ROLES.get(a, ""))
    g["Allocation (%)"] = g.apply(lambda r: _alloc_pct(r["assignee_id"], r["assignee"]), axis=1)
    g["Task Remain"] = (g["total_task"] - g["task_done"]).clip(lower=0)
    g["Bug Remain"]  = (g["total_bug"] - g["bug_done"]).clip(lower=0)
    g["Total Estimate (h)"] = g["total_estimate"].round(1)
    g["Total Actual (h)"]   = g["total_actual"].round(1)
    g["Productivity (%)"]   = g.apply(
        lambda r: _productivity(r["total_actual"], r["total_estimate"]) or "",
        axis=1,
    )
    g["Effort Load"] = g.apply(
        lambda r: _effort_label(r["total_estimate"], r["Allocation (%)"]),
        axis=1,
    )

    g = g.rename(columns={
        "assignee"  : "Member",
        "total_task": "Total Task",
        "task_done" : "Task Done",
        "total_bug" : "Total Bug",
        "bug_done"  : "Bug Done",
    })

    col_order = [
        "Member", "Role", "Allocation (%)", "Effort Load",
        "Total Task", "Task Done", "Task Remain",
        "Total Bug",  "Bug Done",  "Bug Remain",
        "Total Estimate (h)", "Total Actual (h)", "Productivity (%)",
    ]
    return g[col_order].sort_values("Member").reset_index(drop=True)


def _compute_monthly_stats(df: pd.DataFrame, ref_date: str) -> pd.DataFrame:
    """Tổng hợp toàn bộ tháng × assignee. df phải là full-month data (từ get_full_month_data)."""
    if df.empty:
        return pd.DataFrame()

    target_month = pd.Timestamp(ref_date).to_period("M").strftime("%Y-%m")
    sub = df[(df["month"] == target_month) & (df["assignee"] != "Unassigned")].copy()
    if sub.empty:
        return pd.DataFrame()

    sub["_done"] = sub["status"].isin(DONE_STATUSES).astype(int)
    sub["_is_bug"] = (sub["issueType"].str.strip().str.lower() == "bug").astype(int)
    sub["_bug_done"] = (sub["_is_bug"] & sub["_done"]).astype(int)

    g = (
        sub.groupby("assignee")
        .agg(
            assignee_id    =("assigneeId",     "first"),
            total_task     =("id",             "count"),
            task_done      =("_done",          "sum"),
            total_bug      =("_is_bug",        "sum"),
            bug_done       =("_bug_done",      "sum"),
            total_estimate =("estimatedHours", "sum"),
            total_actual   =("actualHours",    "sum"),
        )
        .reset_index()
    )

    g["Role"] = g["assignee"].map(lambda a: MEMBER_ROLES.get(a, ""))
    g["Allocation (%)"] = g.apply(lambda r: _alloc_pct(r["assignee_id"], r["assignee"]), axis=1)
    g["Task Remain"] = (g["total_task"] - g["task_done"]).clip(lower=0)
    g["Bug Remain"]  = (g["total_bug"] - g["bug_done"]).clip(lower=0)
    g["Total Estimate (h)"] = g["total_estimate"].round(1)
    g["Total Actual (h)"]   = g["total_actual"].round(1)
    g["Productivity (%)"]   = g.apply(
        lambda r: _productivity(r["total_actual"], r["total_estimate"]) or "",
        axis=1,
    )
    g["Bug_Rate (%)"] = g.apply(
        lambda r: round(r["total_bug"] / r["total_estimate"] * 100, 1)
        if r["total_estimate"] > 0 else "",
        axis=1,
    )

    g["Tháng"] = target_month
    g = g.rename(columns={
        "assignee"  : "Member",
        "total_task": "Total Task",
        "task_done" : "Task Done",
        "total_bug" : "Total Bug",
        "bug_done"  : "Bug Done",
    })

    col_order = [
        "Tháng", "Member", "Role", "Allocation (%)",
        "Total Task", "Task Done", "Task Remain",
        "Total Bug",  "Bug Done",  "Bug Remain",
        "Total Estimate (h)", "Total Actual (h)", "Productivity (%)",
        "Bug_Rate (%)",
    ]
    return g[col_order].sort_values("Member").reset_index(drop=True)


def _working_days_range(start: pd.Timestamp, n_days: int) -> List[pd.Timestamp]:
    """Tạo danh sách n_days ngày làm việc (bỏ T7, CN) kể từ start."""
    result, d = [], start
    while len(result) < n_days:
        if d.weekday() < 5:   # 0=Mon … 4=Fri
            result.append(d)
        d += timedelta(days=1)
    return result


def _working_days_in_range(start: pd.Timestamp, end: pd.Timestamp) -> List[pd.Timestamp]:
    """Tạo danh sách tất cả ngày làm việc (bỏ T7, CN) trong [start, end] inclusive."""
    result, d = [], start.normalize()
    end_d = end.normalize()
    while d <= end_d:
        if d.weekday() < 5:
            result.append(d)
        d += timedelta(days=1)
    return result


def _count_working_days(start: pd.Timestamp, end: pd.Timestamp) -> int:
    """Đếm số ngày làm việc (T2–T6) trong khoảng [start, end] inclusive."""
    count, d = 0, start
    while d <= end:
        if d.weekday() < 5:
            count += 1
        d += timedelta(days=1)
    return max(1, count)


def _calc_due_from_est(start: pd.Timestamp, estimated_hours: float) -> pd.Timestamp:
    """Tính dueDate ước tính từ startDate + estimatedHours (skip weekend)."""
    import math
    working_days_needed = max(1, math.ceil(estimated_hours / WORKING_HOURS_PER_DAY))
    d, count = start.normalize(), 0
    while True:
        if d.weekday() < 5:
            count += 1
            if count >= working_days_needed:
                return d
        d += timedelta(days=1)


def _calc_start_from_due(due: pd.Timestamp, estimated_hours: float) -> pd.Timestamp:
    """Tính startDate ước tính bằng cách đếm ngược từ dueDate (skip weekend)."""
    import math
    working_days_needed = max(1, math.ceil(estimated_hours / WORKING_HOURS_PER_DAY))
    d, count = due.normalize(), 0   # normalize về midnight
    while True:
        if d.weekday() < 5:
            count += 1
            if count >= working_days_needed:
                return d
        d -= timedelta(days=1)


def _get_action_required_tasks(df: pd.DataFrame) -> pd.DataFrame:
    """
    Gộp tất cả task active cần PM chú ý vào 1 sheet duy nhất.
    Mỗi task có cột 'Loại cảnh báo' liệt kê rõ vấn đề:
      - Thiếu Estimate    : estimatedHours = 0
      - Thiếu Start Date  : startDate trống
      - Thiếu Due Date    : dueDate trống
    Một task có thể có nhiều cảnh báo cùng lúc.
    """
    active = df[
        ~df["status"].isin(DONE_STATUSES) & (df["assignee"] != "Unassigned")
    ].copy()

    def _warnings(row) -> str:
        w = []
        if row["estimatedHours"] == 0:
            w.append("Thiếu Estimate")
        if pd.isna(row["startDate"]):
            w.append("Thiếu Start Date")
        if pd.isna(row["dueDate"]):
            w.append("Thiếu Due Date")
        return " | ".join(w)

    active["Loại cảnh báo"] = active.apply(_warnings, axis=1)
    result = active[active["Loại cảnh báo"] != ""].copy()

    cols = [
        "issueKey", "backlogUrl", "summary", "assignee", "status", "priority",
        "category", "startDate", "dueDate", "estimatedHours", "actualHours",
        "Loại cảnh báo", "updated",
    ]
    return (
        result[[c for c in cols if c in result.columns]]
        .sort_values(["assignee", "Loại cảnh báo"])
        .reset_index(drop=True)
    )


def _debug_availability(df: pd.DataFrame, today: pd.Timestamp) -> None:
    """
    Xuất file CSV debug_availability.csv liệt kê toàn bộ active tasks
    kèm startDate/dueDate sau khi suy luận — để kiểm tra Member_Availability.
    Chỉ ghi file, không ảnh hưởng đến output Excel.
    """
    active = df[
        df["assigneeId"].notna()
        & (~df["status"].isin(DONE_STATUSES))
        & (df["assignee"] != "Unassigned")
    ].copy()

    today_d = today.normalize()
    for col in ("startDate", "dueDate"):
        active[col] = active[col].dt.normalize()

    active["_inferred_start"] = ""
    active["_inferred_due"]   = ""

    for idx, task in active.iterrows():
        has_start = pd.notna(task["startDate"])
        has_due   = pd.notna(task["dueDate"])
        has_est   = task["estimatedHours"] > 0

        if not has_due and has_est:
            start = task["startDate"] if has_start else today_d
            active.at[idx, "dueDate"]         = _calc_due_from_est(start, task["estimatedHours"])
            active.at[idx, "startDate"]       = start
            active.at[idx, "_inferred_due"]   = "YES"
            active.at[idx, "_inferred_start"] = "" if has_start else "YES(today)"

        elif not has_start and has_due and has_est:
            active.at[idx, "startDate"]       = _calc_start_from_due(task["dueDate"], task["estimatedHours"])
            active.at[idx, "_inferred_start"] = "YES(from_due)"

    cols = [
        "issueKey", "assignee", "summary", "status",
        "startDate", "_inferred_start", "dueDate", "_inferred_due",
        "estimatedHours",
    ]
    out = active[[c for c in cols if c in active.columns]].sort_values(["assignee", "startDate"])
    out.to_csv("debug_availability.csv", index=False, encoding="utf-8-sig")
    logger.info("  [DEBUG] debug_availability.csv đã ghi — kiểm tra startDate/dueDate sau suy luận")


def _compute_member_availability(
    df: pd.DataFrame,
    from_date: pd.Timestamp,
    to_date: pd.Timestamp,
) -> pd.DataFrame:
    """
    Bảng số giờ load / ngày của từng member trong tuần báo cáo (Mon–Fri).

    Columns:
      Member | Role | Allocation (%) | Mon MM/DD | Tue MM/DD | Wed MM/DD | Thu MM/DD | Fri MM/DD

    Giá trị mỗi ô = Σ (estimate / duration) của các task overlap ngày đó,
    trong đó duration = số ngày làm việc trong [startDate, dueDate] (min 1).
    """
    if df.empty:
        return pd.DataFrame()

    from_d = from_date.normalize()
    to_d   = to_date.normalize()

    # Lấy 5 ngày làm việc trong tuần báo cáo
    work_days = [d for d in _working_days_in_range(from_d, to_d)]
    if not work_days:
        return pd.DataFrame()

    base = df[df["assigneeId"].notna() & (df["assignee"] != "Unassigned")].copy()
    for col in ("startDate", "dueDate"):
        base[col] = base[col].dt.normalize()

    # Suy ra startDate/dueDate nếu thiếu (dùng lại logic cũ inline)
    for idx, task in base.iterrows():
        has_start = pd.notna(task["startDate"])
        has_due   = pd.notna(task["dueDate"])
        has_est   = task["estimatedHours"] > 0
        if not has_due and has_est:
            start = task["startDate"] if has_start else from_d
            base.at[idx, "dueDate"]   = _calc_due_from_est(start, task["estimatedHours"])
            base.at[idx, "startDate"] = start
        elif not has_start and has_due and has_est:
            base.at[idx, "startDate"] = _calc_start_from_due(task["dueDate"], task["estimatedHours"])
        elif not has_start and not has_due:
            upd = task["updated"].normalize() if pd.notna(task["updated"]) else from_d
            base.at[idx, "startDate"] = upd
            base.at[idx, "dueDate"]   = upd

    base = base[base["startDate"].notna() & base["dueDate"].notna()]
    # Loại các task chưa có estimate — không có giờ để load
    base = base[base["estimatedHours"] > 0]

    members = sorted(m for m in df["assignee"].unique() if m != "Unassigned")
    if not members:
        return pd.DataFrame()

    day_labels = [d.strftime("%a %m/%d") for d in work_days]

    records = []
    for member in members:
        m_id = base.loc[base["assignee"] == member, "assigneeId"].dropna()
        m_id = int(m_id.iloc[0]) if not m_id.empty else None
        row: Dict = {
            "Member": member,
            "Role": MEMBER_ROLES.get(member, ""),
            "Allocation (%)": _alloc_pct(m_id, member),
        }

        m_tasks = base[base["assignee"] == member]
        for day, label in zip(work_days, day_labels):
            busy = m_tasks[(m_tasks["startDate"] <= day) & (m_tasks["dueDate"] >= day)]
            total_load = 0.0
            for _, task in busy.iterrows():
                span = max(1, _count_working_days(task["startDate"], task["dueDate"]))
                total_load += task["estimatedHours"] / span
            row[label] = round(total_load, 1)

        records.append(row)

    cols = ["Member", "Role", "Allocation (%)"] + day_labels
    return pd.DataFrame(records)[cols]


# ============================================================
# 3. PLAN RESOURCE — Lập kế hoạch nhân sự tháng tới
# ============================================================

# Mapping tên team trong need.md → plan team key
_NEED_TEAM_NAME_MAP = {
    "unity":       "Unity",
    "back end":    "BE",
    "backend":     "BE",
    "be":          "BE",
    "fe":          "FE",
    "frontend":    "FE",
    "art":         "Design",
    "design":      "Design",
    "qc":          "QC",
    "gd":          "GD",
    "game design": "GD",
    "blockchain":  "BE",
}


def _parse_need_hours(need_text: str, last_month: str) -> Dict[str, float]:
    """
    Parse need.md và trả về extra hours theo team cho last_month.

    Hỗ trợ 2 loại section:

    1. Section tháng (có scale):
       Tháng N <mô tả>         ← header tháng cụ thể
         TeamName: Xh          ← giờ
       Tháng M keep X% ...     ← tháng sau scale % so với tháng trước

    2. Section fixed (N> header, không có "Tháng N"):
       N> <mô tả>              ← header flat — KHÔNG scale, cộng thẳng vào mọi tháng
         TeamName: Xh

    Ví dụ need.md:
      1> Tháng 3 Livestream
      Unity: 163h
      Tháng 4 keep 50% số trên     ← tháng 4 = 50% × tháng 3
      2> fix bugs, CR              ← flat, cộng thêm vào tháng nào cũng được
      Unity: 154h
    """
    import re as _re

    try:
        target_num = int(last_month.split("-")[1])
    except (IndexError, ValueError):
        return {}

    # ── Pass 1: thu thập sections ──────────────────────────
    month_sections: Dict[int, Dict] = {}   # month_num → {hours, scale}
    flat_hours:     Dict[str, float] = {}  # cộng thẳng, không scale

    cur_num:   Optional[int]   = None
    cur_hours: Dict[str, float] = {}
    cur_scale: Optional[float]  = None
    in_flat:   bool             = False

    def _flush_month():
        nonlocal cur_num, cur_hours, cur_scale
        if cur_num is not None:
            month_sections[cur_num] = {"hours": cur_hours.copy(), "scale": cur_scale}
        cur_num = None; cur_hours = {}; cur_scale = None

    for line in need_text.splitlines():
        s = line.strip()
        if not s:
            continue

        # "N> ..." header — nếu không chứa "Tháng N" → section flat
        numbered = _re.match(r'^\d+>\s*(.*)', s)
        if numbered:
            inner = numbered.group(1).strip()
            month_in_header = _re.search(r'tháng\s+(\d+)', inner, _re.IGNORECASE)
            if month_in_header:
                # Kiểu "1> Tháng 3 ..."  → vẫn là section tháng
                _flush_month()
                in_flat   = False
                cur_num   = int(month_in_header.group(1))
                cur_hours = {}
                cur_scale = None
                pm = _re.search(r'keep\s+(\d+)\s*%', inner, _re.IGNORECASE)
                if pm:
                    cur_scale = float(pm.group(1)) / 100.0
            else:
                # "2> mô tả không có tháng" → section flat
                _flush_month()
                in_flat = True
            continue

        # "Tháng N ..." header (không có "N>" prefix)
        hdr = _re.match(r'tháng\s+(\d+)', s, _re.IGNORECASE)
        if hdr:
            _flush_month()
            in_flat   = False
            cur_num   = int(hdr.group(1))
            cur_hours = {}
            cur_scale = None
            pm = _re.search(r'keep\s+(\d+)\s*%', s, _re.IGNORECASE)
            if pm:
                cur_scale = float(pm.group(1)) / 100.0
            continue

        # "keep X%" độc lập
        pm = _re.search(r'keep\s+(\d+)\s*%', s, _re.IGNORECASE)
        if pm and not in_flat and cur_num is not None:
            cur_scale = float(pm.group(1)) / 100.0
            continue

        # "TeamName: Xh"
        hm = _re.match(r'^(.+?)\s*:\s*(\d+(?:\.\d+)?)\s*h', s, _re.IGNORECASE)
        if hm:
            name = hm.group(1).strip().lower()
            val  = float(hm.group(2))
            for key, team in _NEED_TEAM_NAME_MAP.items():
                if name == key or name.startswith(key) or key in name:
                    if in_flat:
                        flat_hours[team] = flat_hours.get(team, 0.0) + val
                    elif cur_num is not None:
                        cur_hours[team] = cur_hours.get(team, 0.0) + val
                    break

    _flush_month()

    # ── Pass 2: tính kết quả cho target_month ─────────────
    result: Dict[str, float] = {}

    if target_num in month_sections:
        sec   = month_sections[target_num]
        hours = sec["hours"]
        scale = sec["scale"]

        # Section không có giờ nhưng có scale → lấy từ tháng trước
        if not hours and scale is not None and (target_num - 1) in month_sections:
            hours = month_sections[target_num - 1]["hours"]

        if scale is not None and hours:
            result = {t: round(v * scale, 1) for t, v in hours.items()}
        else:
            result = dict(hours)

    # Cộng flat hours (không scale, áp dụng mọi tháng)
    for team, val in flat_hours.items():
        result[team] = round(result.get(team, 0.0) + val, 1)

    return result


def _compute_plan_stats(df: pd.DataFrame, to_date_str: str) -> Dict:
    """
    Tổng hợp dữ liệu tháng cuối + kế hoạch từ plan_resource.json để lập kế hoạch nhân sự.

    Returns dict:
      last_month        : "2026-04"
      next_month        : "2026-05"
      last_month_label  : "April 2026"
      next_month_label  : "May 2026"
      team_stats        : DataFrame — workload thực tế từ Backlog
      member_detail     : DataFrame — chi tiết member từ Backlog
      plan_roster       : DataFrame — kế hoạch từ plan_resource.json
      plan_team_summary : DataFrame — tổng hợp kế hoạch theo team
      need_workload     : DataFrame — extra hours từ need.md
      combined_workload : DataFrame — tổng cộng Backlog + need.md
    """
    to_ts = pd.Timestamp(to_date_str)
    last_period  = to_ts.to_period("M")
    next_period  = last_period + 1

    last_month       = str(last_period)
    next_month       = str(next_period)
    last_month_label = to_ts.strftime("%B %Y")
    next_month_label = (next_period.to_timestamp()).strftime("%B %Y")

    # Filter df for last month
    if "refDate" in df.columns:
        mask = df["refDate"].dt.to_period("M") == last_period
        mdf  = df[mask].copy()
    else:
        mdf = df.copy()

    # ── Planned roster from plan_resource.json ─────────────
    plan_rows = []
    for entry in PLAN_RESOURCE:
        role = entry.get("role", "")
        team = _PLAN_ROLE_TO_TEAM.get(role, "Other")
        plan_rows.append({
            "Team":      team,
            "Role":      role,
            "Full_Name": entry.get("full_name", ""),
            "Account":   entry.get("account", ""),
            "Plan_%":    entry.get("plan", 1.0),
        })
    plan_roster = pd.DataFrame(plan_rows) if plan_rows else pd.DataFrame(
        columns=["Team", "Role", "Full_Name", "Account", "Plan_%"]
    )

    # Team summary from plan_resource
    plan_team_rows = []
    for team in _PLAN_TEAM_GROUPS:
        grp = plan_roster[plan_roster["Team"] == team] if not plan_roster.empty else pd.DataFrame()
        plan_team_rows.append({
            "Team":          team,
            "Planned_HC":    len(grp),
            "Planned_FTE":   round(grp["Plan_%"].sum(), 2) if not grp.empty else 0.0,
            "Members":       ", ".join(grp["Full_Name"].tolist()) if not grp.empty else "",
        })
    plan_team_summary = pd.DataFrame(plan_team_rows)

    # ── Actual workload from Backlog (last month) ──────────
    # Build team → roster mapping from MEMBER_ROLES
    team_roster: Dict[str, List[str]] = {t: [] for t in _PLAN_TEAM_GROUPS}
    for member, role in MEMBER_ROLES.items():
        team = _PLAN_ROLE_TO_TEAM.get(role)
        if team:
            team_roster[team].append(member)

    member_rows = []
    if not mdf.empty and "assignee" in mdf.columns:
        for member, grp in mdf.groupby("assignee"):
            role = MEMBER_ROLES.get(member, "")
            team = _PLAN_ROLE_TO_TEAM.get(role, "Other")
            if team == "Other":
                continue
            member_rows.append({
                "Member":       member,
                "Role":         role,
                "Team":         team,
                "Tasks":        len(grp),
                "Bugs":         int(grp["isBug"].sum()) if "isBug" in grp.columns else 0,
                "Est_Hours":    round(grp["estimatedHours"].sum(), 1) if "estimatedHours" in grp.columns else 0.0,
                "Actual_Hours": round(grp["actualHours"].sum(), 1) if "actualHours" in grp.columns else 0.0,
            })
    member_detail = pd.DataFrame(member_rows) if member_rows else pd.DataFrame(
        columns=["Member", "Role", "Team", "Tasks", "Bugs", "Est_Hours", "Actual_Hours"]
    )

    team_rows = []
    for team in _PLAN_TEAM_GROUPS:
        roster   = team_roster[team]
        team_mdf = mdf[mdf["assignee"].isin(roster)] if not mdf.empty and "assignee" in mdf.columns else pd.DataFrame()
        tasks    = len(team_mdf)
        bugs     = int(team_mdf["isBug"].sum())            if not team_mdf.empty and "isBug"          in team_mdf.columns else 0
        est_h    = round(team_mdf["estimatedHours"].sum(), 1) if not team_mdf.empty and "estimatedHours" in team_mdf.columns else 0.0
        act_h    = round(team_mdf["actualHours"].sum(),    1) if not team_mdf.empty and "actualHours"    in team_mdf.columns else 0.0
        overdue  = int(team_mdf["isOverdue"].sum())        if not team_mdf.empty and "isOverdue"       in team_mdf.columns else 0
        in_data  = team_mdf["assignee"].nunique()          if not team_mdf.empty and "assignee"        in team_mdf.columns else 0
        team_rows.append({
            "Team":            team,
            "Active_In_Month": in_data,
            "Tasks":           tasks,
            "Bugs":            bugs,
            "Est_Hours":       est_h,
            "Actual_Hours":    act_h,
            "Overdue":         overdue,
        })
    team_stats = pd.DataFrame(team_rows)

    # ── need.md extra workload ─────────────────────────────
    need_hours_map = _parse_need_hours(NEED_EXTRA, last_month) if NEED_EXTRA else {}

    need_rows = []
    for team in _PLAN_TEAM_GROUPS:
        need_h = need_hours_map.get(team, 0.0)
        need_rows.append({"Team": team, "Need_Hours": need_h})
    need_workload = pd.DataFrame(need_rows)

    # ── combined workload (Backlog + need.md) ──────────────
    # Working hours capacity per FTE in last_month
    month_start = pd.Timestamp(last_month + "-01")
    month_end   = (month_start + pd.offsets.MonthEnd(0))
    fte_hours   = len(pd.bdate_range(month_start, month_end)) * WORKING_HOURS_PER_DAY

    combined_rows = []
    for _, r in team_stats.iterrows():
        team   = r["Team"]
        est_h  = r["Est_Hours"]
        need_h = need_hours_map.get(team, 0.0)
        total  = round(est_h + need_h, 1)
        fte    = round(total / fte_hours, 2) if fte_hours else 0.0
        combined_rows.append({
            "Team":                team,
            "Est_Hours(Backlog)":  est_h,
            "Need_Hours(need.md)": need_h,
            "Total_Hours":         total,
            f"FTE_Need({fte_hours}h/FTE)": fte,
        })
    combined_workload = pd.DataFrame(combined_rows)

    return {
        "last_month":        last_month,
        "next_month":        next_month,
        "last_month_label":  last_month_label,
        "next_month_label":  next_month_label,
        "team_stats":        team_stats,
        "member_detail":     member_detail,
        "plan_roster":       plan_roster,
        "plan_team_summary": plan_team_summary,
        "need_workload":     need_workload,
        "combined_workload": combined_workload,
        "fte_hours_per_month": fte_hours,
    }


# ============================================================
# 4. OUTPUT — Xuất file Excel
# ============================================================

def export_to_excel(
    df: pd.DataFrame,
    stats: Dict,
    output_file: str,
    plan_data: Optional[Dict] = None,
) -> str:
    """
    Lưu tất cả kết quả vào file .xlsx:
    Dashboard | Monthly_Stats | Member_Availability | Action_Required | Raw_Data | PLAN_Resource*
    * PLAN_Resource chỉ có khi plan_data được truyền vào
    """
    logger.info(f"Exporting to Excel → {output_file}")

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        # Sheet 1: Dashboard (weekly per-member) — startrow=1 dành row 1 cho period header
        dashboard = stats.get("dashboard", pd.DataFrame())
        if not dashboard.empty:
            dashboard.to_excel(writer, sheet_name="Dashboard", index=False, startrow=1)
            logger.info(f"  ✓ Dashboard ({len(dashboard)} members)")
        else:
            pd.DataFrame(columns=[
                "Member", "Role", "Allocation (%)", "Effort Load",
                "Total Task", "Task Done", "Task Remain",
                "Total Bug", "Bug Done", "Bug Remain",
                "Total Estimate (h)", "Total Actual (h)", "Productivity (%)",
            ]).to_excel(writer, sheet_name="Dashboard", index=False, startrow=1)
            logger.info("  ✓ Dashboard (0 members — không có task dueDate trong tuần)")

        # Sheet 2: Monthly Stats
        monthly = stats.get("monthly", pd.DataFrame())
        if not monthly.empty:
            monthly.to_excel(writer, sheet_name="Monthly_Stats", index=False)
            logger.info("  ✓ Monthly_Stats")

        # Sheet 3: Member Availability
        avail = stats.get("availability", pd.DataFrame())
        if not avail.empty:
            avail.to_excel(writer, sheet_name="Member_Availability", index=False)
            logger.info("  ✓ Member_Availability")

        # Sheet 4: Action Required (thiếu estimate / start / due)
        _AR_COLS = ["issueKey", "backlogUrl", "summary", "assignee", "status", "priority",
                    "category", "startDate", "dueDate", "estimatedHours", "actualHours",
                    "Loại cảnh báo", "updated"]
        action_req = stats.get("action_required", pd.DataFrame())
        if not action_req.empty:
            action_req.to_excel(writer, sheet_name="Action_Required", index=False, startrow=1)
            logger.info(f"  ✓ Action_Required ({len(action_req)} tasks)")
        else:
            pd.DataFrame(columns=_AR_COLS).to_excel(writer, sheet_name="Action_Required", index=False, startrow=1)
            logger.info("  ✓ Action_Required (0 tasks — không có vấn đề)")

        # Sheet 5: Raw Data
        if not df.empty:
            raw_cols_drop = ["refDate", "isoWeek", "isoYear", "yearWeek"]
            raw_export = df.drop(columns=raw_cols_drop, errors="ignore")
            raw_export.to_excel(writer, sheet_name="Raw_Data", index=False)
            logger.info("  ✓ Raw_Data")

        # Sheet 6: PLAN_Resource (optional — only when --plan-new-month)
        if plan_data:
            _write_plan_resource_sheet(writer, plan_data)

    # Apply visual formatting after writing
    _format_excel(output_file, stats)

    logger.info(f"File saved: {output_file}")
    return output_file


# ── Formatting helpers ──────────────────────────────────────

_HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
_ALT_ROW_FILL = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
_BUSY_FILL    = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Red
_HALF_FILL    = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")  # Yellow
_FREE_FILL    = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")  # Green
_THIN_BORDER  = Border(
    left  =Side(style="thin", color="CCCCCC"),
    right =Side(style="thin", color="CCCCCC"),
    top   =Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _write_plan_resource_sheet(writer: pd.ExcelWriter, plan_data: Dict) -> None:
    """
    Ghi sheet PLAN_Resource gồm 4 phần:
      1. Title
      2. Kế hoạch nhân sự từ plan_resource.json (plan roster + team summary)
      3. Bảng workload thực tế tháng cuối (từ Backlog)
      4. Chi tiết workload theo member
    """
    wb = writer.book
    ws = wb.create_sheet("PLAN_Resource")

    last_lbl       = plan_data.get("last_month_label",  "")
    next_lbl       = plan_data.get("next_month_label",  "")
    team_df        = plan_data.get("team_stats",         pd.DataFrame())
    member_df      = plan_data.get("member_detail",      pd.DataFrame())
    plan_roster_df = plan_data.get("plan_roster",        pd.DataFrame())
    plan_team_df   = plan_data.get("plan_team_summary",  pd.DataFrame())

    # ── Style constants ────────────────────────────────────
    title_fill   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_font   = Font(color="FFFFFF", bold=True, size=14)
    section_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    section_font = Font(color="FFFFFF", bold=True, size=11)
    hdr_fill     = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    hdr_font     = Font(bold=True, size=10, color="1F4E79")
    alt_fill     = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    thin         = Border(
        left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC"),
    )
    team_colors = {
        "Unity":  PatternFill(start_color="E9D7F5", end_color="E9D7F5", fill_type="solid"),
        "BE":     PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
        "FE":     PatternFill(start_color="D9EDF7", end_color="D9EDF7", fill_type="solid"),
        "QC":     PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "Design": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "GD":     PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    }

    row = 1

    # ── Row 1: Title ───────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    c = ws.cell(row=row, column=1,
                value=f"  PLAN Resource — {next_lbl}  (dựa trên dữ liệu {last_lbl})")
    c.fill = title_fill; c.font = title_font
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 30
    row += 2

    # ── Section 1: Planned Roster (plan_resource.json) ─────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    c = ws.cell(row=row, column=1, value=f"  📋 Kế hoạch nhân sự hiện tại (plan_resource.json) — dùng cho tháng {next_lbl}")
    c.fill = section_fill; c.font = section_font
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    # Plan team summary (left side: cols 1-4)
    if not plan_team_df.empty:
        pt_labels = {"Team": "Team", "Planned_HC": "Headcount plan", "Planned_FTE": "FTE plan", "Members": "Thành viên"}
        pt_cols = list(plan_team_df.columns)
        for c_idx, col in enumerate(pt_cols, 1):
            c = ws.cell(row=row, column=c_idx, value=pt_labels.get(col, col))
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal="center"); c.border = thin
        row += 1
        for i, (_, r_data) in enumerate(plan_team_df.iterrows()):
            team_name = r_data.get("Team", "")
            base_fill = team_colors.get(team_name, alt_fill if i % 2 == 0 else PatternFill())
            for c_idx, col in enumerate(pt_cols, 1):
                c = ws.cell(row=row, column=c_idx, value=r_data[col])
                c.fill = base_fill; c.border = thin
                c.alignment = Alignment(horizontal="left" if c_idx in (1, 4) else "center",
                                        wrap_text=(c_idx == 4))
            row += 1
        row += 1

    need_df     = plan_data.get("need_workload",     pd.DataFrame())
    combined_df = plan_data.get("combined_workload", pd.DataFrame())
    fte_cap     = plan_data.get("fte_hours_per_month", 0)

    need_src_label = f"need.md (nguồn: {_NEED_MD_PATH})" if NEED_EXTRA else "need.md (trống)"
    _need_md_fill  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    _need_md_font  = Font(size=10, color="7F6000")
    _total_fill    = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    _total_hdr_fill= PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    _total_hdr_font= Font(color="FFFFFF", bold=True, size=10)

    def _write_generic_table(ws, start_row, section_title, df_table, col_label_map,
                             section_bg, section_fg_font, hdr_bg, hdr_fg, row_bg_fn,
                             n_merge_cols=8):
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=n_merge_cols)
        ch = ws.cell(row=start_row, column=1, value=section_title)
        ch.fill = section_bg; ch.font = section_fg_font
        ch.alignment = Alignment(vertical="center")
        ws.row_dimensions[start_row].height = 22
        start_row += 1
        if df_table.empty:
            ws.cell(row=start_row, column=1, value="(Không có dữ liệu)")
            return start_row + 2
        cols = list(df_table.columns)
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=start_row, column=ci, value=col_label_map.get(col, col))
            cell.fill = hdr_bg; cell.font = hdr_fg
            cell.alignment = Alignment(horizontal="center"); cell.border = thin
        start_row += 1
        for i, (_, r_data) in enumerate(df_table.iterrows()):
            for ci, col in enumerate(cols, 1):
                cell = ws.cell(row=start_row, column=ci, value=r_data[col])
                cell.fill = row_bg_fn(r_data.get("Team", ""), i)
                cell.border = thin
                cell.alignment = Alignment(horizontal="left" if ci == 1 else "center")
            start_row += 1
        return start_row + 1

    def _team_row_fill(team_name, idx):
        return team_colors.get(team_name, alt_fill if idx % 2 == 0 else PatternFill())

    # ── Section 2: Workload thực tế từ Backlog ─────────────
    backlog_col_labels = {
        "Team":            "Team",
        "Active_In_Month": "Có task trong tháng",
        "Tasks":           "Tổng Tasks",
        "Bugs":            "Bugs",
        "Est_Hours":       "Est Hours",
        "Actual_Hours":    "Actual Hours",
        "Overdue":         "Overdue",
    }
    row = _write_generic_table(ws, row,
        f"  📊 Workload thực tế từ Backlog — {last_lbl}",
        team_df, backlog_col_labels,
        section_fill, section_font, hdr_fill, hdr_font, _team_row_fill)

    # ── Section 3: Workload need từ need.md ────────────────
    need_col_labels = {
        "Team":       "Team",
        "Need_Hours": f"Need Hours ({need_src_label})",
    }
    # Show raw need.md lines as a sub-note before the table
    if NEED_EXTRA:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cn = ws.cell(row=row, column=1, value=f"  📌 Nội dung need.md")
        cn.fill = _need_md_fill; cn.font = Font(bold=True, size=10, color="7F6000")
        cn.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1
        for line in NEED_EXTRA.splitlines():
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            cn = ws.cell(row=row, column=1, value=line)
            cn.fill = _need_md_fill; cn.font = _need_md_font
            cn.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        row += 1

    row = _write_generic_table(ws, row,
        f"  📊 Workload need từ need.md — {last_lbl}",
        need_df, need_col_labels,
        section_fill, section_font, hdr_fill, hdr_font, _team_row_fill)

    # ── Section 4: Workload tổng cần ───────────────────────
    fte_col_key = f"FTE_Need({fte_cap}h/FTE)" if fte_cap else "FTE_Need"
    combined_col_labels = {
        "Team":                "Team",
        "Est_Hours(Backlog)":  "Est Hours (Backlog)",
        "Need_Hours(need.md)": "Need Hours (need.md)",
        "Total_Hours":         "Total Hours",
        fte_col_key:           f"FTE cần ({fte_cap}h/FTE)",
    }
    row = _write_generic_table(ws, row,
        f"  📊 Workload tổng cần (Backlog + need.md) — {last_lbl}",
        combined_df, combined_col_labels,
        _total_hdr_fill, _total_hdr_font,
        PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid"),
        Font(bold=True, size=10, color="375623"),
        lambda t, i: _total_fill)

    # ── Section 5: Member Detail Table ────────────────────
    if not member_df.empty:
        m_col_labels = {
            "Member":       "Member",
            "Role":         "Role",
            "Team":         "Team",
            "Tasks":        "Tasks",
            "Bugs":         "Bugs",
            "Est_Hours":    "Est Hours",
            "Actual_Hours": "Actual Hours",
        }
        row = _write_generic_table(ws, row,
            f"  👤 Chi tiết workload member — {last_lbl}",
            member_df, m_col_labels,
            section_fill, section_font, hdr_fill, hdr_font, _team_row_fill)

    # ── Column widths ──────────────────────────────────────
    col_widths = [30, 16, 16, 12, 10, 12, 14, 12, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    logger.info(f"  ✓ PLAN_Resource ({last_lbl} → plan {next_lbl})")


_PERF_GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_PERF_GREEN_FONT  = Font(color="276221", bold=True)
_PERF_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_PERF_YELLOW_FONT = Font(color="9C6500", bold=True)
_PERF_RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_PERF_RED_FONT    = Font(color="9C0006", bold=True)


def _format_excel(filepath: str, stats: Dict) -> None:
    """Áp dụng style/màu sắc toàn bộ file."""
    wb = load_workbook(filepath)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Dashboard & Action_Required có description ở row 1 → header ở row 2
        header_row = 2 if sheet_name in ("Action_Required", "Dashboard") else 1

        # Auto column width
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            max_len = max(
                (len(str(c.value)) for c in col_cells if c.value is not None),
                default=10,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

        # Header row style
        ws.row_dimensions[header_row].height = 30
        ws.freeze_panes = f"A{header_row + 1}"
        for cell in ws[header_row]:
            cell.fill      = _HEADER_FILL
            cell.font      = _HEADER_FONT
            cell.border    = _THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Data rows
        for idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
            for cell in row:
                cell.border    = _THIN_BORDER
                cell.alignment = Alignment(vertical="center")
                if idx % 2 == 0:
                    cell.fill = _ALT_ROW_FILL

        # Hyperlink cho cột backlogUrl (áp dụng mọi sheet có cột này)
        _apply_hyperlinks(ws)

        if sheet_name == "Dashboard":
            _write_dashboard_period_header(ws, stats.get("period"))
            _style_dashboard_sheet(ws, header_row=header_row)
            _add_legend(ws, [
                ("Productivity > 100%", "Estimate ÷ Actual × 100",
                 "🟢 Hoàn thành nhanh hơn estimate", "ok"),
                ("Productivity 80–100%", "Estimate ÷ Actual × 100",
                 "🟡 Làm đúng effort dự kiến", "warn"),
                ("Productivity < 80%",  "Estimate ÷ Actual × 100",
                 "🔴 Tốn nhiều effort hơn estimate", "bad"),
                ("Effort Load = Còn dư", "Estimate < 87.5% × Alloc-hours (100% = 40h/tuần)",
                 "🟡 Còn dư capacity — có thể assign thêm", "warn"),
                ("Effort Load = Đủ",     "Estimate ≈ 87.5–100% × Alloc-hours (100% = 40h)",
                 "🟢 Cân bằng, đúng chuẩn effort tuần", "ok"),
                ("Effort Load = Quá tải", "Estimate > Alloc-hours (100% = 40h/tuần)",
                 "🔴 Quá tải — cần chia bớt task", "bad"),
            ], "Chú thích — Dashboard")

        if sheet_name == "Member_Availability":
            _style_availability_sheet(ws)
            ws.freeze_panes = "D2"  # Freeze Member / Role / Allocation
            _add_legend(ws, [
                ("Ô 🟢 Xanh",   "Load < 4h",    "Còn nhiều capacity trong ngày", "ok"),
                ("Ô 🟡 Vàng",   "Load 4h – 7h", "Bận nhưng còn chỗ", "warn"),
                ("Ô 🔴 Đỏ",     "Load > 7h",    "Quá tải trong ngày (vượt 7h/ngày làm việc)", "bad"),
                ("Cách tính load", "Σ (estimate ÷ duration)",
                 "duration = số ngày làm việc (Mon–Fri) từ start → due, tối thiểu 1", ""),
            ], "Chú thích — Member Availability")

        if sheet_name == "Monthly_Stats":
            _style_month_stats_sheet(ws)
            _add_legend(ws, [
                ("Productivity > 100%", "Estimate ÷ Actual × 100",
                 "🟢 Hoàn thành nhanh hơn estimate", "ok"),
                ("Productivity 80–100%", "Estimate ÷ Actual × 100",
                 "🟡 Làm đúng effort dự kiến", "warn"),
                ("Productivity < 80%",  "Estimate ÷ Actual × 100",
                 "🔴 Tốn nhiều effort hơn estimate", "bad"),
                ("Bug_Rate < 10%",  "Total Bug ÷ Total Estimate × 100",
                 "🟢 Chất lượng tốt", "ok"),
                ("Bug_Rate ≥ 10%",  "Total Bug ÷ Total Estimate × 100",
                 "🔴 Nhiều bug so với effort — cần review chất lượng", "bad"),
            ], "Chú thích — Monthly Stats")

        if sheet_name == "Action_Required":
            _write_action_required_description(ws)
            _style_missing_schedule_sheet(ws, header_row=header_row)
            _add_legend(ws, [
                ("Thiếu Estimate",   "estimatedHours = 0",
                 "🔴 Task chưa có ước lượng giờ — không tính vào Effort Load / Availability", "bad"),
                ("Thiếu Start Date", "startDate = null",
                 "🟡 Không xác định được khi nào bắt đầu — script sẽ suy luận ngược từ Due", "warn"),
                ("Thiếu Due Date",   "dueDate = null",
                 "🟡 Không track được deadline — script sẽ suy luận từ Start + Estimate", "warn"),
                ("Ô vàng nhạt",      "Ô trống trong Start / Due / Estimate",
                 "Highlight ô cụ thể bị thiếu để PM dễ nhìn", ""),
            ], "Chú thích — Action Required")

    wb.save(filepath)


def _write_dashboard_period_header(ws, period) -> None:
    """Ghi row 1 của Dashboard: 'Báo cáo hiệu suất tuần: DD/MM/YYYY → DD/MM/YYYY (Tuần W##/YYYY)'."""
    if not period:
        return
    from_date, to_date = period
    from_ts = pd.Timestamp(from_date)
    to_ts   = pd.Timestamp(to_date)
    iso = from_ts.isocalendar()
    label = (
        f"📊 Báo cáo hiệu suất tuần: "
        f"{from_ts.strftime('%d/%m/%Y')} → {to_ts.strftime('%d/%m/%Y')} "
        f"(Tuần W{int(iso.week):02d}/{iso.year})"
    )

    ncols = max(ws.max_column, 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=label)
    cell.fill      = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    cell.font      = Font(bold=True, size=12, color="FFFFFF")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26


def _write_action_required_description(ws) -> None:
    """Ghi description ở row 1 của Action_Required."""
    ncols = max(ws.max_column, 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=(
        "📋 Danh sách backlog cần PM sửa — thiếu Estimate / Start Date / Due Date. "
        "Bổ sung các trường thiếu để báo cáo hiệu suất chính xác."
    ))
    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    cell.font = Font(bold=True, size=10, color="7F6000")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24


_EFFORT_FILLS = {
    "Còn dư":  (PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
                Font(color="9C6500", bold=True)),
    "Đủ":      (PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                Font(color="276221", bold=True)),
    "Quá tải": (PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                Font(color="9C0006", bold=True)),
}


def _color_productivity(cell) -> None:
    """Productivity = Est/Actual*100 — >100 xanh (nhanh hơn), 80-100 vàng (đúng), <80 đỏ (tốn effort)."""
    try:
        v = float(cell.value) if cell.value not in (None, "") else None
    except (ValueError, TypeError):
        v = None
    if v is None:
        return
    if v > 100:
        cell.fill = _PERF_GREEN_FILL; cell.font = _PERF_GREEN_FONT
    elif v >= 80:
        cell.fill = _PERF_YELLOW_FILL; cell.font = _PERF_YELLOW_FONT
    else:
        cell.fill = _PERF_RED_FILL; cell.font = _PERF_RED_FONT


def _style_dashboard_sheet(ws, header_row: int = 1) -> None:
    """Tô Productivity (%) + Effort Load trong Dashboard."""
    header = {cell.value: cell.column for cell in ws[header_row] if cell.value}
    prod_col   = header.get("Productivity (%)")
    effort_col = header.get("Effort Load")
    for row in ws.iter_rows(min_row=header_row + 1):
        r = row[0].row
        if prod_col:
            _color_productivity(ws.cell(row=r, column=prod_col))
        if effort_col:
            cell = ws.cell(row=r, column=effort_col)
            style = _EFFORT_FILLS.get(str(cell.value or "").strip())
            if style:
                cell.fill, cell.font = style


def _style_month_stats_sheet(ws, header_row: int = 1) -> None:
    """Tô Productivity + Bug_Rate trong Monthly_Stats."""
    header = {cell.value: cell.column for cell in ws[header_row] if cell.value}
    prod_col = header.get("Productivity (%)")
    bug_col  = header.get("Bug_Rate (%)")

    for row in ws.iter_rows(min_row=header_row + 1):
        r = row[0].row
        if prod_col:
            _color_productivity(ws.cell(row=r, column=prod_col))
        if bug_col:
            cell = ws.cell(row=r, column=bug_col)
            try:
                v = float(cell.value) if cell.value not in (None, "") else None
            except (ValueError, TypeError):
                v = None
            if v is None:
                continue
            if v < 10:
                cell.fill = _PERF_GREEN_FILL; cell.font = _PERF_GREEN_FONT
            else:
                cell.fill = _PERF_RED_FILL; cell.font = _PERF_RED_FONT


def _style_availability_sheet(ws) -> None:
    """
    Tô các cột ngày trong Member_Availability theo mức load.
    Load = Σ (estimate / duration) — số giờ tải trong ngày.
      = 0h    : trắng (không có task)
      < 4h    : xanh (rảnh, có thể nhận thêm)
      4h–7h   : vàng (bận, còn chỗ)
      > 7h    : đỏ (quá tải trong ngày)
    """
    for row in ws.iter_rows(min_row=2):
        # 3 cột đầu = Member / Role / Allocation (%)
        for cell in row[3:]:
            try:
                val = float(cell.value) if cell.value is not None else None
            except (ValueError, TypeError):
                continue
            if val is None:
                continue
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.value = f"{val}h"
            if val <= 0:
                continue
            if val < 4:
                cell.fill = _FREE_FILL
            elif val <= 7:
                cell.fill = _HALF_FILL
            else:
                cell.fill = _BUSY_FILL


_LEGEND_HEADER_FILL  = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
_LEGEND_HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
_LEGEND_LABEL_FILL   = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_LEGEND_LABEL_FONT   = Font(bold=True, size=9, color="1F4E79")
_LEGEND_VALUE_FILL   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_LEGEND_GOOD_FILL    = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")   # Xanh nhạt
_LEGEND_WARN_FILL    = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # Vàng nhạt
_LEGEND_BAD_FILL     = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")   # Đỏ nhạt


def _add_legend(ws, legend_data: list, title: str) -> None:
    """
    Vẽ bảng chú thích chỉ số sang bên phải vùng dữ liệu (cách 2 cột).
    legend_data: list of (chỉ số, công thức, cách đọc, mức)
    """
    start_col = ws.max_column + 2
    start_row = 1

    # Tiêu đề bảng legend (span 3 cột)
    title_cell = ws.cell(row=start_row, column=start_col, value=f"📖  {title}")
    title_cell.fill      = _LEGEND_HEADER_FILL
    title_cell.font      = Font(color="FFFFFF", bold=True, size=11)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[start_row].height = max(ws.row_dimensions[start_row].height or 15, 22)

    # Header cột
    hdr_row = start_row + 1
    for col_offset, label in enumerate(["Chỉ số", "Công thức / Nguồn gốc", "Cách đọc"], start=0):
        c = ws.cell(row=hdr_row, column=start_col + col_offset, value=label)
        c.fill      = _LEGEND_HEADER_FILL
        c.font      = _LEGEND_HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _THIN_BORDER
    ws.row_dimensions[hdr_row].height = 18

    # Dữ liệu từng chỉ số
    LEVEL_FILLS = {"ok": _LEGEND_GOOD_FILL, "warn": _LEGEND_WARN_FILL,
                   "bad": _LEGEND_BAD_FILL, "": _LEGEND_VALUE_FILL}

    for i, (metric, formula, reading, level) in enumerate(legend_data):
        r = hdr_row + 1 + i
        row_fill = LEVEL_FILLS.get(level, _LEGEND_VALUE_FILL)

        cells_vals = [metric, formula, reading]
        for col_offset, val in enumerate(cells_vals):
            c = ws.cell(row=r, column=start_col + col_offset, value=val)
            c.fill      = _LEGEND_LABEL_FILL if col_offset == 0 else row_fill
            c.font      = _LEGEND_LABEL_FONT if col_offset == 0 else Font(size=9)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border    = _THIN_BORDER
        ws.row_dimensions[r].height = 28

    # Column widths cho vùng legend
    ws.column_dimensions[get_column_letter(start_col)].width     = 22
    ws.column_dimensions[get_column_letter(start_col + 1)].width = 30
    ws.column_dimensions[get_column_letter(start_col + 2)].width = 45


def _apply_hyperlinks(ws) -> None:
    """
    Tìm cột 'backlogUrl' trong sheet và chuyển thành hyperlink có thể click.
    Cell hiển thị issueKey, click mở Backlog.
    """
    header = {cell.value: cell.column for cell in ws[1] if cell.value}
    url_col    = header.get("backlogUrl")
    key_col    = header.get("issueKey")

    if not url_col:
        return

    link_font = Font(color="0563C1", underline="single", bold=False)

    for row in ws.iter_rows(min_row=2):
        row_idx  = row[0].row
        url_cell = ws.cell(row=row_idx, column=url_col)
        url      = url_cell.value

        if not url:
            continue

        # Lấy issueKey từ cột riêng (nếu có) để dùng làm display text
        display = url
        if key_col:
            key_val = ws.cell(row=row_idx, column=key_col).value
            if key_val:
                display = str(key_val)

        url_cell.value     = display
        url_cell.hyperlink = url
        url_cell.font      = link_font


def _style_missing_schedule_sheet(ws, header_row: int = 1) -> None:
    """Highlight ô trống trong startDate/dueDate/estimatedHours và cột Loại cảnh báo."""
    _EMPTY_FILL   = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Vàng nhạt
    _MISSING_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")  # Đỏ đậm
    _MISSING_FONT = Font(color="FFFFFF", bold=True)

    header = {cell.value: cell.column for cell in ws[header_row] if cell.value}
    watch_cols = {
        header.get("startDate"),
        header.get("dueDate"),
        header.get("estimatedHours"),
    } - {None}
    missing_col = header.get("Loại cảnh báo")

    for row in ws.iter_rows(min_row=header_row + 1):
        row_idx = row[0].row
        for col_idx in watch_cols:
            cell = ws.cell(row=row_idx, column=col_idx)
            val  = cell.value
            if val is None or str(val).strip() in ("", "0", "0.0"):
                cell.fill = _EMPTY_FILL

        if missing_col:
            cell = ws.cell(row=row_idx, column=missing_col)
            if cell.value:
                cell.fill = _MISSING_FILL
                cell.font = _MISSING_FONT


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Backlog Statistics Report Generator",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python backlog_report.py
  python backlog_report.py --from-date 2025-01-01 --to-date 2025-03-31
  python backlog_report.py --from-date 2025-01-01 --to-date 2025-03-31 --output Q1_report.xlsx
        """,
    )
    default_to   = datetime.now()
    default_from = default_to - timedelta(days=30)

    parser.add_argument(
        "--from-date",
        default=default_from.strftime("%Y-%m-%d"),
        help="Ngày bắt đầu (YYYY-MM-DD) — mặc định: 30 ngày trước",
    )
    parser.add_argument(
        "--to-date",
        default=default_to.strftime("%Y-%m-%d"),
        help="Ngày kết thúc (YYYY-MM-DD) — mặc định: hôm nay",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Tên file xuất ra (mặc định: {projectKey}_{fromYYYYMMDD}_{toYYYYMMDD}.xlsx)",
    )
    parser.add_argument(
        "--plan-new-month",
        action="store_true",
        default=False,
        help=(
            "Tạo thêm sheet PLAN_Resource: phân tích tháng cuối trong khoảng --to-date "
            "và đề xuất kế hoạch nhân sự tháng tiếp theo (có buffer +30%%)"
        ),
    )
    parser.add_argument(
        "--data-file",
        default=None,
        help="Đường dẫn tới file JSON cache issues (bỏ qua Backlog API, dùng dữ liệu local)",
    )

    args = parser.parse_args()

    # Tự sinh tên file mặc định theo format {projectKey}_YYYYMMDD_YYYYMMDD.xlsx
    if not args.output:
        proj_key = get_project_key()
        from_yyyymmdd = pd.Timestamp(args.from_date).strftime("%Y%m%d")
        to_yyyymmdd   = pd.Timestamp(args.to_date).strftime("%Y%m%d")
        args.output = f"{proj_key}_{from_yyyymmdd}_{to_yyyymmdd}.xlsx"

    SEP = "=" * 60
    logger.info(SEP)
    logger.info("  BACKLOG STATISTICS REPORT GENERATOR")
    logger.info(SEP)
    logger.info(f"  Period : {args.from_date} → {args.to_date}")
    logger.info(f"  Output : {args.output}")
    if args.plan_new_month:
        next_m = (pd.Timestamp(args.to_date).to_period("M") + 1).to_timestamp().strftime("%B %Y")
        logger.info(f"  Plan   : ENABLED — sẽ tạo sheet PLAN_Resource cho {next_m}")
    logger.info(SEP)

    # ── Step 1: Extract ──────────────────────────────────────
    if args.data_file:
        logger.info(f"Đọc dữ liệu từ file cache: {args.data_file}")
        with open(args.data_file, encoding="utf-8") as _cache_f:
            _raw = json.load(_cache_f)
        df = _parse_issues(_raw)
        from_d = pd.Timestamp(args.from_date).normalize()
        to_d   = pd.Timestamp(args.to_date).normalize()
        # Filter theo dueDate trong khoảng — đây là tiêu chí chính
        df = df[df["dueDate"].notna() & (df["dueDate"].dt.normalize() >= from_d) & (df["dueDate"].dt.normalize() <= to_d)].reset_index(drop=True)
        logger.info(f"Issues có dueDate trong [{args.from_date} → {args.to_date}]: {len(df)}")
    else:
        df = get_backlog_data(args.from_date, args.to_date)
    if df.empty:
        logger.error("Không có dữ liệu. Kết thúc.")
        return

    # ── Tự fill backlog_id còn trống trong plan_resource.json từ dữ liệu Backlog ───
    _autofill_plan_resource_ids(df)

    # ── Step 1b: Fetch full month data for Monthly_Stats ────────
    month_df: Optional[pd.DataFrame] = None
    if not args.data_file:
        month_df = get_full_month_data(args.from_date)

    # ── Step 2: Process ──────────────────────────────────────
    stats = process_stats(df, args.from_date, args.to_date, month_df=month_df)

    # ── Step 3: Plan Resource (optional) ─────────────────────
    plan_data: Optional[Dict] = None
    if args.plan_new_month:
        logger.info("Đang tính toán plan resource...")
        plan_data = _compute_plan_stats(df, args.to_date)

    # ── Step 4: Export ───────────────────────────────────────
    export_to_excel(df, stats, args.output, plan_data=plan_data)

    logger.info("")
    logger.info(f"✅  Hoàn thành! File: {args.output}")
    logger.info("")


if __name__ == "__main__":
    main()
